"""One-shot parser that turns the client's Vendor + Customer Excels
into TypeScript array literals we can paste into prisma/seed.ts.

Run from anywhere; outputs both blocks to stdout:
  python scripts/parse-imports.py > /tmp/import.ts
"""
import openpyxl
import re
import json


def ts_str(s):
    """JSON-encode then strip surrounding quotes — handles all escapes."""
    if s is None:
        return "null"
    return json.dumps(s, ensure_ascii=False)


def parse_vendors():
    wb = openpyxl.load_workbook(
        "C:/Users/laksh/Downloads/Vendor details.xlsx",
        data_only=True,
        read_only=True,
    )
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:  # skip header
        if not r or not r[0]:
            continue
        name = (str(r[0]) or "").strip()
        if not name:
            continue
        phone = (str(r[1]).strip() if r[1] else None) or None
        address = (str(r[2]).strip() if r[2] else None) or None
        gst = (str(r[3]).strip() if r[3] else None) or None
        email = (str(r[4]).strip() if r[4] else None) or None
        bank = (str(r[5]).strip() if r[5] else None) or None
        msme = (str(r[6]).strip() if r[6] else None) or None

        unregistered = bool(gst) and gst.upper() in ("UNREGISTERED", "0")
        gst_norm = None if unregistered or not gst else gst.replace(" ", "")
        state = "29"
        if gst_norm and len(gst_norm) >= 2 and gst_norm[:2].isdigit():
            state = gst_norm[:2]

        msme_bool = bool(
            msme
            and msme.upper() not in ("UNREGISTERED", "0", "")
            and msme.strip()
        )

        notes_parts = []
        if bank:
            notes_parts.append("Bank: " + re.sub(r"\s+", " ", bank).strip())
        if msme and msme_bool:
            notes_parts.append("MSME ref: " + msme.strip())
        if unregistered:
            notes_parts.append("Unregistered (no GST)")
        notes = "\n".join(notes_parts) if notes_parts else None

        out.append(
            {
                "name": name,
                "gstin": gst_norm,
                "stateCode": state,
                "phone": phone,
                "email": email if (email and "@" in email) else None,
                "address": address,
                "msme": msme_bool,
                "notes": notes,
            }
        )
    return out


def parse_customers():
    wb = openpyxl.load_workbook(
        "C:/Users/laksh/Downloads/CX billing Details.xlsx",
        data_only=True,
        read_only=True,
    )
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))

    # Group into blocks separated by blank rows. Each block is one customer.
    blocks = []
    cur = []
    for r in rows:
        cell = (str(r[1]).strip() if r and len(r) > 1 and r[1] else "")
        if cell:
            cur.append(cell)
        else:
            if cur:
                blocks.append(cur)
                cur = []
    if cur:
        blocks.append(cur)

    out = []
    for b in blocks:
        # Find GST line, vendor code, payment terms.
        gst = None
        vendor_code = None
        payment_terms = None
        addr_lines = []
        for line in b:
            low = line.lower()
            if "gst" in low and ":" in line:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    gst = parts[1].strip()
            elif "vendor code" in low and ":" in line:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    vendor_code = parts[1].strip()
            elif (
                "payment terms" in low
                or "pay terms" in low
                or "payment term" in low
            ) and ":" in line:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    payment_terms = parts[1].strip()
            elif (
                line.startswith("29A")
                or line.startswith("33A")
                or re.match(r"^[0-9]{2}[A-Z]{5}", line)
            ) and len(line) >= 15:
                # Standalone GSTIN line (no "GST:" prefix)
                gst = line.strip()
            else:
                addr_lines.append(line)

        if not addr_lines:
            continue

        # Name = best-guess: usually the most institution-y line.
        # Pick the second line if first is "The Chairman" / "The Chair" / "The Director".
        title_words = {"the chairman", "the chair", "the director", "the registrar"}
        name = None
        for ln in addr_lines:
            if ln.lower().strip() not in title_words:
                # Drop generic prefixes.
                name = ln.strip()
                break
        if not name:
            name = addr_lines[0]

        # Build billingAddress = full block joined
        billing = "\n".join(addr_lines)

        # GSTIN normalize
        gst_norm = None
        if gst:
            gst_norm = re.sub(r"\s+", "", gst.upper())
            # Reject if it doesn't look like a GSTIN
            if not re.match(r"^[0-9]{2}[A-Z]{5}", gst_norm):
                gst_norm = None
        state = "29"
        if gst_norm:
            state = gst_norm[:2]

        # Credit days from payment terms text — extract first integer.
        credit_days = 0
        if payment_terms:
            m = re.search(r"(\d+)", payment_terms)
            if m:
                credit_days = int(m.group(1))

        notes_parts = []
        if vendor_code:
            notes_parts.append(
                "Customer's vendor code for us: " + vendor_code
            )
        if payment_terms:
            notes_parts.append("Payment terms: " + payment_terms)

        out.append(
            {
                "name": name,
                "gstin": gst_norm,
                "stateCode": state,
                "billingAddress": billing,
                "creditDays": credit_days,
                "notes": "\n".join(notes_parts) if notes_parts else None,
            }
        )
    return out


def emit_vendors(vendors):
    print(
        f"// {len(vendors)} vendors imported from "
        f"Vendor details.xlsx — Phase 6 seed addendum."
    )
    print("const vendorRecords: VendorSeed[] = [")
    for v in vendors:
        print(
            "  {{ name: {n}, gstin: {g}, stateCode: {s}, phone: {p}, email: {e}, address: {a}, msme: {m}, notes: {note} }},".format(
                n=ts_str(v["name"]),
                g=ts_str(v["gstin"]),
                s=ts_str(v["stateCode"]),
                p=ts_str(v["phone"]),
                e=ts_str(v["email"]),
                a=ts_str(v["address"]),
                m="true" if v["msme"] else "false",
                note=ts_str(v["notes"]),
            )
        )
    print("];")
    print()


def emit_customers(customers):
    print(
        f"// {len(customers)} customers imported from "
        f"CX billing Details.xlsx — Phase 6 seed addendum."
    )
    print("const customerRecords: CustomerSeed[] = [")
    for c in customers:
        print(
            "  {{ name: {n}, gstin: {g}, stateCode: {s}, billingAddress: {a}, creditDays: {d}, notes: {note} }},".format(
                n=ts_str(c["name"]),
                g=ts_str(c["gstin"]),
                s=ts_str(c["stateCode"]),
                a=ts_str(c["billingAddress"]),
                d=c["creditDays"],
                note=ts_str(c["notes"]),
            )
        )
    print("];")


if __name__ == "__main__":
    vendors = parse_vendors()
    customers = parse_customers()
    emit_vendors(vendors)
    emit_customers(customers)
